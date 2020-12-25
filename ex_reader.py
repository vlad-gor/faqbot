import openpyxl
from openpyxl import load_workbook

wb = load_workbook('base/FAQ_CityStore.xlsx')

sheet = wb['FAQ WEB']

def get_catagories(sheet, line):
    c1 = sheet.cell(row=line, column=1).value
    c2 = sheet.cell(row=line, column=2).value
    c3 = sheet.cell(row=line, column=3).value
    return (c1,c2,c3)

def get_ans(sheet, line):
    c1 = sheet.cell(row=line, column=1).value
    c2 = sheet.cell(row=line, column=2).value
    c3 = sheet.cell(row=line, column=3).value
    return (c1,c2,c3)

def get_values(sheet, line):
    c1 = sheet.cell(row=line, column=1).value
    c2 = sheet.cell(row=line, column=2).value
    c3 = sheet.cell(row=line, column=3).value
    return (c1,c2,c3)


class Question():

    def __init__(self, category, answer, translate):
        self.category = category
        self.answer = answer
        self.translate = translate

    def ru(self):
        return f"{self.category[0]} | {self.answer[0]} | {self.translate[0]}"

    def en(self):
        return f"{self.category[1]} | {self.answer[1]} | {self.translate[1]}"

    def dt(self):
        return f"{self.category[2]} | {self.answer[2]} | {self.translate[2]}"

    def __repr__(self):
        return f"{self.category[0]} | {self.answer[0]} | {self.translate[0]}"

    def get_answer(self, lang):
        if lang == 'ru':
            return self.ru().split('|')[2].strip()
        elif lang == 'en':
            return self.en().split('|')[2].strip()
        elif lang == 'dt':
            return self.dt().split('|')[2].strip()

    def get_question(self, lang):
        if lang == 'ru':
            return self.ru().split('|')[1].strip()
        elif lang == 'en':
            return self.en().split('|')[1].strip()
        elif lang == 'dt':
            return self.dt().split('|')[1].strip()


class Question_List():

    def __init__(self):
        cats = [3, 22, 31, 46, 59, 70, 75, 81]
        self.ql = list()

        for ci, cat in enumerate(cats):
            try:
                for i in range(cat+1, cats[ci+1], 2):
                    self.ql.append(
                        Question(
                            get_catagories(sheet, cat),
                            get_ans(sheet, i),
                            get_values(sheet, i+1)
                        ))
            except:
                pass

