from fuzzywuzzy import fuzz
import pymorphy2
import spacy
import openpyxl
from openpyxl import load_workbook
import logging

log = logging.getLogger(__name__)

wb = load_workbook('base/FAQ_CityStore.xlsx')
sheet = wb['FAQ WEB']

morph = pymorphy2.MorphAnalyzer()

nlp_en = spacy.load("en_core_web_sm")
nlp_de = spacy.load("de_core_news_sm")

def get_catagories(sheet, line):
    c1 = sheet.cell(row=line, column=1).value
    c2 = sheet.cell(row=line, column=2).value
    c3 = sheet.cell(row=line, column=3).value
    return (c1,c2,c3)

def get_ans(sheet, line):
    c1 = sheet.cell(row=line, column=1).value
    c2 = sheet.cell(row=line, column=2).value
    c3 = sheet.cell(row=line, column=3).value
    return (c1,c2,c3)

def get_values(sheet, line):
    c1 = sheet.cell(row=line, column=1).value
    c2 = sheet.cell(row=line, column=2).value
    c3 = sheet.cell(row=line, column=3).value
    return (c1,c2,c3)


class Question():

    def __init__(self, category, answer, translate):
        self.category = category
        self.answer = answer
        self.translate = translate

    def en(self):
        return f"{self.category[1]} | {self.answer[1]} | {self.translate[1]}"

    def dt(self):
        return f"{self.category[2]} | {self.answer[2]} | {self.translate[2]}"

    def __repr__(self):
        return f"{self.category[0]} | {self.answer[0]} | {self.translate[0]}"

    def get_answer(self, lang):
        if lang == 'English':
            return self.en().split('|')[2].strip()
        elif lang == 'German':
            return self.dt().split('|')[2].strip()

    def get_question(self, lang):
        if lang == 'English':
            return self.en().split('|')[1].strip()
        elif lang == 'German':
            return self.dt().split('|')[1].strip()


class Analyzer():

    def __init__(self):
        cats = [3, 22, 31, 46, 59, 70, 75, 81]
        self.ql = list()
        for ci, cat in enumerate(cats):
            try:
                for i in range(cat+1, cats[ci+1], 2):
                    self.ql.append(Question(get_catagories(sheet, cat),
                            get_ans(sheet, i),get_values(sheet, i+1)))
            except:
                pass 
        self.nlp_en_quest = [nlp_en(q.get_question('English')) for q in self.ql]
        self.nlp_de_quest = [nlp_de(q.get_question('German')) for q in self.ql]

    def classify_question_morphy(self, message, lang):
        '''Функция поиска ответа на вопрос'''
        text = ' '.join(morph.parse(word)[0].normal_form for word in message.text.split())
        scores = list()
        for quest in self.ql:
            norm_question = ' '.join(morph.parse(word)[0].normal_form for word in quest.get_question(lang).split())
            scores.append(fuzz.token_sort_ratio(norm_question.lower(), text.lower()))
        log.info(scores)
        answer = self.ql[scores.index(max(scores))].get_answer(lang)
        return answer

    def classify_question_spacy(self, message, lang):
        scores = list()
        if lang == 'English':
            nlp_mes = nlp_en(message.text)
            scores = [round(nlp_mes.similarity(nlp_q), 4) for nlp_q in self.nlp_en_quest]
        elif lang == 'German':
            nlp_mes = nlp_de(message.text)
            scores = [round(nlp_mes.similarity(nlp_q), 4) for nlp_q in self.nlp_de_quest]
        log.info(scores)
        answer = self.ql[scores.index(max(scores))].get_answer(lang)
        return answer

analyst = Analyzer()